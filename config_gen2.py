import sys
import os
import csv
import socket
import struct
import pprint
import ipaddress
from jinja2 import Template
from zipfile import ZipFile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from datetime import date

# Feature Requests
# V1 Working with generating multiple device templates with one go on user inputs
# V1.25 All Versions of Config for all the Juniper Device Types
# V1.5 Working with Gitlab to pull latest template / upload configs
# V1.75 Custom Entries to adjust the config / Review before Printing
# V2 Has a GUI Front end for easy drop down box's and check boxes
# V3 Works with Sky Enterprises and uploads configs to Sky for Zero Touch Provisioning
# V4 Auto Updates Solarwinds & Atlas (Monitoring, Inventory)
# V5 Works with back end Database for IP's & Management





#Fure Function to pull latest Templates from Github



#Open Templates
#dir = os.chdir(r'opt/NetDevOps/Config_gen2')
#cpath = os.getcwd()
#ef_single = open(cpath + '/ef_single.conf', 'r')
#ef_cluster = open(cpath + '/ef_single.conf', 'r')
#cs_vc = open(cpath + '/cs_vc.conf', 'r')
#as_vc = open(cpath + '/as_vc.conf', 'r')
#he = open(cpath + '/he.conf', 'r')
dir2 = os.chdir(r'/opt/NetDevOps/Config_gen2')
cpath2 = os.getcwd()
#os.chdir(dir2)
jinsrx = open(cpath2 + '/ef_template.j2', 'r')
jinex = open(cpath2 + '/ex_template.j2', 'r')
jinhe = open(cpath2 + '/he_template.j2', 'r')

#ef_single_srx240 = (open(cpath + '/ef_single_srx20'))

# Original Code for Example
#fhub = open(cpath + '/template/hub.conf', 'r')
#fspoke = open(cpath + '/template/spoke.conf', 'r')

# Read the Data from the files
#FILEDATAEF_SINGLE = ef_single.read()
#FILEDATAEF_CLUSTER = ef_cluster.read()
#FILEDATACS_VC = cs_vc.read()
#FILEDATAAS_VC = as_vc.read()
#FILEDATAHE = he.read()
FILEDATAJINSRX = jinsrx.read()
FILEDATAJINEX = jinex.read()
FILEDATAJINHE = jinhe.read()
#FILEDATAEFS240 = ef_single_srx240.read()

#ef_single.close()
#ef_cluster.close()
#cs_vc.close()
#as_vc.close()
#he.close()
jinsrx.close()
jinex.close()
jinhe.close()
#ef_single_srx240.close()


#Original Code for Example
#filedataHub = fhub.read()
#filedataSpoke = fspoke.read()

def parseHostname(hostname):
    hostname_fields = hostname.split('-')
    sub = 'xxx'
    site = 'xxxxxxxxxxxx'
    site_tier = 0
    device_type = 'xx'
    device_serial = '01'
    if len(hostname_fields) > 1:
        sub = hostname_fields[0]
        site = hostname_fields[1]
    if len(hostname_fields) > 2:
        site_tier = int(hostname_fields[2][0:1])
        device_type = hostname_fields[2][1:3]
        device_serial = hostname_fields[2][3:5]
    return sub, site, site_tier, device_type, device_serial

# Function for Passing Arguments in from PHP (Web Front End)
def passinput(data):
    hostname = data['hostname']
    sub, site, tier, dtype, dserial = parseHostname(hostname)

    model = data['Model']
    core = data['Core']
    switch = data['Switch']
    EF_Type = data['EF_Type']
    VC_Core = data['VC_Core']
    vc = data['VC']
    access = 'No'
    dhcp = data['dhcp']
    st_tun2 = data['st_tun2']

    # Get Input from User for Config Data

    bgp_asn = int(data['bgp_asn'])

    # Add remove first char of ASN
    st_tun = str(bgp_asn)[1:5]
    #st_tun2 = data['st_tun2']

    site_mailing_addr = data['site_mailing_address']

    # Site Super net checker
    subnet = {}
    site_supernet = data['site_supernet'].split('/')
    try:
        socket.inet_aton(site_supernet[0])
        if len(site_supernet) == 2 and int(site_supernet[1]) in range(0, 33):
            if tier == 1:
                subnet = subnet_cont_tier1(site_supernet)
            elif tier in (2, 3):
                subnet = subnet_const_tier3(site_supernet)
    except socket.error:
        print('Invlaid supernet %s, requires x.x.x.x/x\n' % ('/'.join(site_supernet)))
    # end of Supernet checker

    wan1_isp_name = data['wan1_isp_name']
    wan1_isp_ckt_id = data['wan1_isp_ckid']
    wan1_isp_support_num = data['wan1_isp_support_num']

    # Condition for ensuring input has a suffix
#    wan1_isp_up = '0'
#    suffix = ['b', 'k', 'm', 'g']
#    while wan1_isp_up[-1:] not in suffix:
    wan1_isp_up = data['wan1_isp_up'].lower()

#    wan1_isp_down = '0'
#    while wan1_isp_down[-1:] not in suffix:
    wan1_isp_down = data['wan1_isp_down'].lower()

    wan1_isp_ip = data['wan1_isp_ip']
    wan1_isp_gw = data['wan1_isp_gw']

    site_tun1_ip = data['site_tun1_ip']
    atl_tun1_ip = data['atl_tun1_ip']

    site_dal_tun1_ip = data['site_dal_tun1_ip']
    dal_tun1_ip = data['dal_tun1_ip']


    # PreDefines variables based on other Variables
    preshared_key = str.upper(sub) + "!psk0987"
    cs_name = sub + "-" + site + '-' + str(tier) + 'CS01' 
    ef_hostname = sub + '-' + site + '-' + str(tier) + dtype + dserial 
    # Part of the Supernet /31 for tier 1
    # Access Switches + 2 Links (Maybe just put 10)
    ae_count = '10'
    ikehost = sub + '-' + site + '-' + 'ISP1'
    ikehost2 = sub + '-' + site + '-' + 'ISP2'

    ## Defaults
#    wan2_isp_name = "RESERVED"
#    wan2_isp_ckt_id = "RESERVED"
#    wan2_isp_support_num = "RESERVED"
#    wan2_isp_up = '100m'
#    wan2_isp_down = '100m'
#    wan2_isp_ip = '1.1.1.2/30'
#    wan2_isp_gw = '1.1.1.1'
#    vc_n0_sn = "RESERVED"
#    vc_n1_sn = "RESERVED"
    site_atl_tun2_ip = data['site_tun2_ip']
    atl_tun2_ip = data['atl_tun2_ip']
    site_dal_tun2_ip = data['site_dal_tun2_ip']
    dal_tun2_ip = data['dal_tun2_ip']

    cs_n0_sn = data['cs_Serial_n0']
    cs_n1_sn = data['cs_Serial_n1']
    cs_n2_sn = data['cs_Serial_n2']
    cs_n3_sn = data['cs_Serial_n3']

    vc_n0_sn = data['vc_Serial_n0']
    vc_n1_sn = data['vc_Serial_n1']
    vc_n2_sn = data['vc_Serial_n2']
    vc_n3_sn = data['vc_Serial_n3']

    wan2_isp_name = data['wan2_isp_name']
    wan2_isp_ckt_id = data['wan2_isp_ckid']
    wan2_isp_support_num = data['wan2_isp_support_num']
    wan2_isp_up = data['wan2_isp_up']
    wan2_isp_down = data['wan2_isp_down']
    wan2_isp_ip = data['wan2_isp_ip']
    wan2_isp_gw = data['wan2_isp_gw']

    replace_values = {}
    replace_values['<VL300_IP>'] = subnet['vlan300']['ip']
    replace_values['<VL300_BM>'] = subnet['vlan300']['cidr']
    replace_values['<VL300_NW>'] = subnet['vlan300']['net']
    replace_values['<VL300_DHCP_LOW>'] = subnet['vlan300']['dhcplow']
    replace_values['<VL300_DHCP_HIGH>'] = subnet['vlan300']['dhcphigh']
    replace_values['<VL310_IP>'] = subnet['vlan310']['ip']
    replace_values['<VL310_BM>'] = subnet['vlan310']['cidr']
    replace_values['<VL310_NW>'] = subnet['vlan310']['net']
    replace_values['<VL310_DHCP_LOW>'] = subnet['vlan310']['dhcplow']
    replace_values['<VL310_DHCP_HIGH>'] = subnet['vlan310']['dhcphigh']
    replace_values['<VL320_IP>'] = subnet['vlan320']['ip']
    replace_values['<VL320_BM>'] = subnet['vlan320']['cidr']
    replace_values['<VL320_NW>'] = subnet['vlan320']['net']
    replace_values['<VL320_DHCP_LOW>'] = subnet['vlan320']['dhcplow']
    replace_values['<VL320_DHCP_HIGH>'] = subnet['vlan320']['dhcphigh']
    replace_values['<VL330_IP>'] = subnet['vlan330']['ip']
    replace_values['<VL330_BM>'] = subnet['vlan330']['cidr']
    replace_values['<VL330_NW>'] = subnet['vlan330']['net']
    replace_values['<VL330_DHCP_LOW>'] = subnet['vlan330']['dhcplow']
    replace_values['<VL330_DHCP_HIGH>'] = subnet['vlan330']['dhcphigh']
    replace_values['<VL311_IP>'] = subnet['vlan311']['ip']
    replace_values['<VL311_BM>'] = subnet['vlan311']['cidr']
    replace_values['<VL311_NW>'] = subnet['vlan311']['net']
    replace_values['<VL311_DHCP_LOW>'] = subnet['vlan311']['dhcplow']
    replace_values['<VL311_DHCP_HIGH>'] = subnet['vlan311']['dhcphigh']
    replace_values['<VL901_EF01>'] = subnet['vlan901']['ip']
    replace_values['<VL901_BM>'] = subnet['vlan901']['cidr']
    replace_values['<VL901_NW>'] = subnet['vlan901']['net']
    replace_values['<VL901_CS01>'] = subnet['vlan901']['gw']
    replace_values['<VL901_DHCP_LOW>'] = subnet['vlan330']['dhcplow']
    replace_values['<VL901_DHCP_HIGH>'] = subnet['vlan330']['dhcphigh']
    replace_values['<HOSTNAME>'] = hostname
    replace_values['<EF_LOOPBACK>'] = subnet['vlan901']['ef_loopback']
    replace_values['<CS_LOOPBACK>'] = subnet['vlan901']['cs_loopback']
    replace_values['<AE_COUNT>'] = str(ae_count)
    replace_values['<PRESHARED_KEY>'] = preshared_key
    replace_values['<CS_NAME>'] = cs_name
    replace_values['<WAN1_ISPNAME>'] = wan1_isp_name
    replace_values['<WAN1_CKT_ID>'] = wan1_isp_ckt_id
    replace_values['<WAN1_SUPPORT_NO>'] = wan1_isp_support_num
    replace_values['<WAN1_UP>'] = wan1_isp_up
    replace_values['<WAN1_DOWN>'] = wan1_isp_down
    replace_values['<WAN1_IP>'] = wan1_isp_ip
    replace_values['<WAN1_GW>'] = wan1_isp_gw
    replace_values['<BGP_ASN>'] = bgp_asn
    replace_values['<ST0>'] = st_tun
    replace_values['<SITE_ADDRESS>'] = site_mailing_addr
    replace_values['<ATL_TUN1_IP>'] = site_tun1_ip
    replace_values['<ATL_TUN1_PEER>'] = atl_tun1_ip
    replace_values['<DAL_TUN1_IP>'] = site_dal_tun1_ip
    replace_values['<DAL_TUN1_PEER>'] = dal_tun1_ip
    replace_values['<SITE_SUPERNET>'] = site_supernet[0] + '/' + site_supernet[1]
    replace_values['<SUB>'] = sub
    replace_values['<SITE>'] = site
    replace_values['<TIER>'] = tier
    replace_values['<DEVICE_TYPE>'] = dtype
    replace_values['<DEVICE_SERIAL>'] = dserial
    replace_values['<SWITCH>'] = switch
    replace_values['<EF01_HOSTNAME>'] = ef_hostname
    replace_values['<CS01_HOSTNAME>'] = cs_name
    replace_values['<TUN1_ST>'] = st_tun
    replace_values['<MODEL>'] = model
    replace_values['<CORE>'] = core
    replace_values['<EF_TYPE>'] = EF_Type
    replace_values['<VC_CORE>'] = int(VC_Core)
    replace_values['<VC>'] = int(vc)
    replace_values['<ACCESS>'] = access
    replace_values['<CORE_PORTS>'] = int(data['Core'][-3:-1])
    replace_values['<ACCESS_PORTS>'] = int(data['Switch'][-3:-1])
    replace_values['<CS_N0_SN>'] = cs_n0_sn
    replace_values['<CS_N1_SN>'] = cs_n1_sn
    replace_values['<CS_N2_SN>'] = cs_n2_sn
    replace_values['<CS_N3_SN>'] = cs_n3_sn
    replace_values['<DHCP>'] = dhcp 
    replace_values['<IKEHOST>'] = ikehost
    replace_values['<IKEHOST2>'] = ikehost2
    accessip = subnet['vlan300']['ip']
    ip = ipaddress.ip_address(accessip) + 1

    replace_values['<ACCESSIP>'] = ip

    #defaults
    replace_values['<WAN2_ISPNAME>'] = wan2_isp_name
    replace_values['<WAN2_CKT_ID>'] = wan2_isp_ckt_id
    replace_values['<WAN2_SUPPORT_NO>'] = wan2_isp_support_num
    replace_values['<WAN2_UP>'] = wan2_isp_up
    replace_values['<WAN2_DOWN>'] = wan2_isp_down
    replace_values['<WAN2_IP>'] = wan2_isp_ip
    replace_values['<WAN2_GW>'] = wan2_isp_gw

    replace_values['<VC_N0_SN>'] = vc_n0_sn
    replace_values['<VC_N1_SN>'] = vc_n1_sn
    replace_values['<VC_N2_SN>'] = vc_n2_sn
    replace_values['<VC_N3_SN>'] = vc_n3_sn

    replace_values['<ATL_TUN2_IP>'] = site_atl_tun2_ip
    replace_values['<ATL_TUN2_PEER>'] = atl_tun2_ip
    replace_values['<DAL_TUN2_IP>'] = site_dal_tun2_ip
    replace_values['<DAL_TUN2_PEER>'] = dal_tun2_ip
    replace_values['<TUN2_ST>'] = st_tun2

    data = replace_values

    data['<CORE_PORTS>'] = data['<CORE_PORTS>'] - 1
    data['<ACCESS_PORTS>'] = data['<ACCESS_PORTS>'] -1
    data['<CORE_PORTS2>'] = data['<CORE_PORTS>'] -1
    data['<ACCESS_PORTS2>'] = data['<ACCESS_PORTS>'] -1
    data['<VC1>'] = data['<VC>'] - 1


    asbuilt(data)
# Run SRX Config to Jinja2
    jinsrx(data)
# Run EX Core Config to Jinja2
    data['<LOOPBACK>'] = data['<CS_LOOPBACK>']
    hostname = data['<HOSTNAME>']
    core_switch_hostname = hostname[0:len(hostname) - 4] + 'CS01' 
    data['<HOSTNAME>'] = core_switch_hostname
    jinex(data)
# Run EX VC Config to Jinja2
    if data['<VC>'] is not 0:
        data['<LOOPBACK>'] = data['<CS_LOOPBACK>']
        data['<ACCESS>'] = 'Yes'
        hostname = data['<HOSTNAME>']
        core_switch_hostname = hostname[0:len(hostname) - 4] + 'AS01' 
        data['<HOSTNAME>'] = core_switch_hostname
        jinex(data)
# Run Head End Config to Jinja2
    hostname = data['<HOSTNAME>']
    core_switch_hostname = hostname[0:len(hostname) - 4] + 'HE'
    data['<HOSTNAME>'] = core_switch_hostname
    jinhe(data)
    hostname = data['<HOSTNAME>']
    zipped = zipfile(str(tier), sub, site, dtype, dserial)
    return zipped



def jinsrx(data):
    global FILEDATAJINSRX
    subdir = '/opt/NetDevOps/configs/'
    path = '/opt/NetDevOps/Config_gen2'
    myfile = open(path + '/ef_template.j2')
    #myfile = FILEDATAJINSRX
    myfile_contents = myfile.read()
    myfile.close()
    template = Template(myfile_contents)
#    print(template.render(data=data))


    hostname = filename = data['<HOSTNAME>']


    try:	
        os.stat(subdir)
    except:
        os.mkdir(subdir)

    with open(subdir + filename + '.conf', 'w') as f:
        f.write(template.render(data=data))
    f.close()
    

    print('Saved as {0}/{1}.conf \n'.format(subdir, hostname))

    return


def jinex(data):
    global FILEDATAJINEX
    subdir = '/opt/NetDevOps/configs/'
    path = '/opt/NetDevOps/Config_gen2'
    myfile2 = open(path + '/ex_template.j2')
    #myfile2 = FILEDATAJINEX
    myfile_contents2 = myfile2.read()
    myfile2.close()
    template2 = Template(myfile_contents2)
#    print(template2.render(data=data))

    hostname = filename = data['<HOSTNAME>']

    try:
        os.stat(subdir)
    except:
        os.mkdir(subdir)
    with open(subdir + filename + '.conf', 'w') as f:
        f.write(template2.render(data=data))
    f.close()

    print('Saved as {0}/{1}.conf \n'.format(subdir, hostname))
    return
# Create Head End Config
#    head_end_Conf(replace_values)

def jinhe(data):
    global FILEDATAJINHE
    subdir = '/opt/NetDevOps/configs/'
    path = '/opt/NetDevOps/Config_gen2'
    myfile = open(path + '/he_template.j2')

    isp1 = data['<WAN1_IP>']
    data['<WAN1_IP>'] = isp1[0:len(isp1) - 3]

    isp2 = data['<WAN2_IP>']
    data['<WAN2_IP>'] = isp2[0:len(isp2) - 3]

    #myfile = FILEDATAJINHE
    myfile_contents = myfile.read()
    myfile.close()
    template = Template(myfile_contents)
#    print(template.render(data=data))

    hostname = filename = data['<HOSTNAME>']

    try:
        os.stat(subdir)
    except:
        os.mkdir(subdir)

    with open(subdir + filename + '.conf', 'w') as f:
        f.write(template.render(data=data))
    f.close()

    print('Saved as {0}/{1}.conf \n'.format(subdir, hostname))


    return

def zipfile(tier, sub, site, dtype, dserial):
    #going to zip some files
    # create a ZipFile object
    filenamed = sub + '-' + site
    path = '/opt/NetDevOps/configs/'
    #os.system('cp ' + '/opt/NetDevOps/Config_gen2/Asbuilt_Template_2.0.xlsx ' + path + filenamed + '.xlsx')
    os.chdir(path)
    zipObj = ZipFile(filenamed + '.zip', 'w')
    zipped = path + filenamed + '.zip'
    # Add multiple files to the zip
    zipObj.write(filenamed + '-' + tier + 'HE' + '.conf')
    zipObj.write(filenamed + '-' + tier + 'CS01' + '.conf')
    zipObj.write(filenamed + '-' + tier + 'EF01' + '.conf')
    zipObj.write(filenamed + '.xlsx') 
    if os.path.exists(filenamed + '-' + tier + 'AS01' + '.conf'):
        zipObj.write(filenamed + '-' + tier + 'AS01' + '.conf')
    else:
        pass	
    # close the Zip File
    zipObj.close()
    if os.path.exists(filenamed + '-' + tier + 'HE' + '.conf'):
    	os.remove(filenamed + '-' + tier + 'HE' + '.conf')
    	os.remove(filenamed + '-' + tier + 'CS01' + '.conf')
    	os.remove(filenamed + '-' + tier + 'EF01' + '.conf')
    	if os.path.exists(filenamed + '-' + tier + 'AS01' + '.conf'):
        	os.remove(filenamed + '-' + tier + 'AS01' + '.conf')
    	else:
        	pass
    else:
    	pass
    os.remove(filenamed + '.xlsx')

    return zipped

def asbuilt(data):
    hostname = data['<HOSTNAME>']
    name = hostname[0:len(hostname) - 6]
    today = date.today()
    tier = data['<TIER>']
    st1 = data['<ST0>']

    path = '/opt/NetDevOps/Config_gen2/'
    # Load Workbook and Edit
    wb = load_workbook(path + 'Asbuilt_Template_2.0.xlsx')
    ws = wb['Summary']
    ws['B1'] = name
    ws['H1'] = today
    ws['B20'] = today
    ws['E1'] = 'Tier' + str(tier)
    ws['H9'] = data['<WAN1_DOWN>']
    ws['H10'] = data['<WAN1_CKT_ID>']
    ws['H11'] = data['<WAN2_DOWN>']
    ws['H12'] = data['<WAN2_CKT_ID>']
    ws['I9'] = data['<WAN1_ISPNAME>']
    ws['I11'] = data['<WAN2_ISPNAME>']
    ws['D29'] = data['<CS_N0_SN>']
    ws['D30'] = data['<CS_N1_SN>']
    ws['D31'] = data['<CS_N2_SN>']
    ws['D32'] = data['<CS_N3_SN>']
    ws['D33'] = data['<VC_N0_SN>']
    ws['D34'] = data['<VC_N1_SN>']
    ws['D35'] = data['<VC_N2_SN>']
    ws['D36'] = data['<VC_N3_SN>']

    # Solve for Data Center Images for HL Design
    path = '/opt/NetDevOps/Config_gen2/img/'
    img = Image(path + 'datacenter.jpg')
    ws = wb['Data Center']
    ws.add_image(img, 'B2')

    # Solve for Rack Design
    path = '/opt/NetDevOps/Config_gen2/img/'
    img = Image(path + 'rack.jpg')
    ws = wb['Rack Design']
    ws.add_image(img, 'A2')

    # Solve for IP Assignments
    ws = wb['IP Assignments']
    ws['A1'] = data['<SUB>']
    ws['A2'] = data['<SUB>']
    ws['B1'] = data['<SITE>']
    ws['B2'] = data['<SITE>']
    ws['C1'] = data['<TIER>']
    ws['C2'] = data['<TIER>']
    ws['D1'] = data['<BGP_ASN>']
    ws['D2'] = data['<BGP_ASN>']
    ws['E1'] = data['<ST0>']
    ws['E2'] = int(st1) + 1
    ws['F2'] = data['<ATL_TUN1_IP>']
    ws['G2'] = data['<ATL_TUN1_PEER>']
    ws['H2'] = data['<DAL_TUN1_IP>']
    ws['I2'] = data['<DAL_TUN1_PEER>']
    ws['J2'] = data['<EF_LOOPBACK>']
    ws['K2'] = data['<CS_LOOPBACK>']
    ws['L2'] = data['<SITE_SUPERNET>']
    ws['M2'] = data['<WAN1_IP>']
    ws['N2'] = data['<WAN1_GW>']
    ws['F3'] = data['<ATL_TUN2_IP>']
    ws['G3'] = data['<ATL_TUN2_PEER>']
    ws['H3'] = data['<DAL_TUN2_IP>']
    ws['I3'] = data['<DAL_TUN2_PEER>']
    ws['J3'] = data['<EF_LOOPBACK>']
    ws['K3'] = data['<CS_LOOPBACK>']
    ws['L3'] = data['<SITE_SUPERNET>']
    ws['M3'] = data['<WAN2_IP>']
    ws['M3'] = data['<WAN2_GW>']

    # Solve for Tier Images for HL Design
    path = '/opt/NetDevOps/Config_gen2/img/'
    if data['<TIER>'] == '1':
        img = Image(path + 'Tier1.jpg')
    elif data['<TIER>'] == '2':
        img = Image(path + 'Tier2.jpg')
    else:
        img = Image(path + 'Tier3.jpg')

    ws = wb['HLDesign']
    ws.add_image(img, 'B2')

    path = '/opt/NetDevOps/configs/'
    wb.save(path + name + '.xlsx')
    return
#Function to get input from user to build config
# Based on Tier 1, 2, 3 It has some static ability but also can be custom per user request
def getinput():
    print ("Welcome to Config_Gen 1.0 \n")
    hostname = input('Full Hostname for Firewall (EXAMPLE: GSU-SACRAMENTO-3EF01)')
    sub, site, tier, dtype, dserial = parseHostname(hostname)

    num_switches = 0
    while num_switches not in range (1,10):
        num_switches = int(input("How many Switches at this site: "))

# Get Input from User for Config Data

    bgp_asn = 0
    while bgp_asn not in range(64512,65535):
        bgp_asn = int(input('What is BGP ASN: (64512 - 65534) '))

# Add remove first char of ASN
    st_tun = str(bgp_asn)[1:5]

    site_mailing_addr = input('What is the sites mailing address: ')

# Site Super net checker
    while True:
        site_supernet = input('Site supernet? (x.x.x.x/x)\n').split('/')
        try:
           socket.inet_aton(site_supernet[0])
           if len(site_supernet) == 2 and int(site_supernet[1]) in range(0,33):
                if tier == 1:
                    subnet = subnet_cont_tier1(site_supernet)
                    break
                elif tier in (2,3):
                    subnet = subnet_const_tier3(site_supernet)
                    break
        except socket.error:
            print('Invlaid supernet %s, requires x.x.x.x/x\n' % ('/'.join(site_supernet)))
# end of Supernet checker

    wan1_isp_name = input('WAN 1 ISP Name (Comcast, Centurylink, ETC : ')
    wan1_isp_ckt_id = input('WAN 1 ISP CKT ID: ')
    wan1_isp_support_num = input('WAN 1 ISP Support Number: ')

# Condition for ensuring input has a suffix
    wan1_isp_up = '0'
    suffix = ['b', 'k', 'm', 'g']
    while wan1_isp_up[-1:] not in suffix:
        wan1_isp_up = input('WAN 1 ISP Upload Speed (provide suffix (b, k, m, g, etc.): ').lower()

    wan1_isp_down = '0'
    while wan1_isp_down[-1:] not in suffix:
       wan1_isp_down = input('WAN 1 ISP Download Speed (provide suffix (b, k, m, g, etc.): ').lower()

    wan1_isp_ip = input('WAN 1 ISP IP x.x.x.x/xx : ' )
    wan1_isp_gw = input('WAN 1 ISP GW x.x.x.x : ')

    site_tun1_ip = input('What is the Site Tunnel IP x.x.x.x : ')
    atl_tun1_ip = input('What is the ATL Tunnel IP x.x.x.x : ')

    site_dal_tun1_ip = input('what is the Site to Dal Tunnel IP x.x.x.x : ')
    dal_tun1_ip = input('What is the Dal Tunnel IP x.x.x.x : ')

    cs_uplink_n0 = input('On what Interface does the EF connect to on CS-N0 (EXM ge-0/0/46) : ')
    cs_uplink_n1 = input('On what Interface does the EF connect to on CS-N1 (EXM ge-0/0/46) : ')

# PreDefines variables based on other Variables
    preshared_key = str.upper(sub) + "!psk0987"
    cs_name = sub + "-" + site + '-' + str(tier) + 'CS01' 
    ef_hostname = sub + '-' + site + '-' + str(tier) + dtype + dserial 
# Part of the Supernet /31 for tier 1
# Access Switches + 2 Links (Maybe just put 10)
    ae_count = '10'

## Defaults
    wan2_isp_name = "RESERVED"
    wan2_isp_ckt_id = "RESERVED"
    wan2_isp_support_num = "RESERVED"
    wan2_isp_up = '100m'
    wan2_isp_down = '100m'
    wan2_isp_ip = '1.1.1.2/30'
    wan2_isp_gw = '1.1.1.1'
    vc_n0_sn = "RESERVED"
    vc_n1_sn = "RESERVED"
    site_atl_tun2_ip = '1.1.1.2/30'
    atl_tun2_ip = '1.1.1.1'
    site_dal_tun2_ip = '1.1.1.2/30'
    dal_tun2_ip = '1.1.1.1'
    vc_n0_sn = '0'
    vc_n1_sn = '0'

# All Tokens will go here
    replace_values = {}
    replace_values['<VL300_IP>'] = subnet['vlan300']['ip']
    replace_values['<VL300_BM>'] = subnet['vlan300']['cidr']
    replace_values['<VL300_NW>'] = subnet['vlan300']['net']
    replace_values['<VL300_DHCP_LOW>'] = subnet['vlan300']['dhcplow']
    replace_values['<VL300_DHCP_HIGH>'] = subnet['vlan300']['dhcphigh']
    replace_values['<VL310_IP>'] = subnet['vlan310']['ip']
    replace_values['<VL310_BM>'] = subnet['vlan310']['cidr']
    replace_values['<VL310_NW>'] = subnet['vlan310']['net']
    replace_values['<VL310_DHCP_LOW>'] = subnet['vlan310']['dhcplow']
    replace_values['<VL310_DHCP_HIGH>'] = subnet['vlan310']['dhcphigh']
    replace_values['<VL320_IP>'] = subnet['vlan320']['ip']
    replace_values['<VL320_BM>'] = subnet['vlan320']['cidr']
    replace_values['<VL320_NW>'] = subnet['vlan320']['net']
    replace_values['<VL320_DHCP_LOW>'] = subnet['vlan320']['dhcplow']
    replace_values['<VL320_DHCP_HIGH>'] = subnet['vlan320']['dhcphigh']
    replace_values['<VL330_IP>'] = subnet['vlan330']['ip']
    replace_values['<VL330_BM>'] = subnet['vlan330']['cidr']
    replace_values['<VL330_NW>'] = subnet['vlan330']['net']
    replace_values['<VL330_DHCP_LOW>'] = subnet['vlan330']['dhcplow']
    replace_values['<VL330_DHCP_HIGH>'] = subnet['vlan330']['dhcphigh']
    replace_values['<VL311_IP>'] = subnet['vlan311']['ip']
    replace_values['<VL311_BM>'] = subnet['vlan311']['cidr']
    replace_values['<VL311_NW>'] = subnet['vlan311']['net']
    replace_values['<VL311_DHCP_LOW>'] = subnet['vlan311']['dhcplow']
    replace_values['<VL311_DHCP_HIGH>'] = subnet['vlan311']['dhcphigh']
    replace_values['<VL901_EF01>'] = subnet['vlan901']['ip']
    replace_values['<VL901_BM>'] = subnet['vlan901']['cidr']
    replace_values['<VL901_NW>'] = subnet['vlan901']['net']
    replace_values['<VL901_CS01>'] = subnet['vlan901']['gw']
    replace_values['<VL901_DHCP_LOW>'] = subnet['vlan330']['dhcplow']
    replace_values['<VL901_DHCP_HIGH>'] = subnet['vlan330']['dhcphigh']
    replace_values['<HOSTNAME>'] = hostname
    replace_values['<EF_LOOPBACK>'] = subnet['vlan901']['ef_loopback']
    replace_values['<CS_LOOPBACK>'] = subnet['vlan901']['cs_loopback']
    replace_values['<AE_COUNT>'] = str(ae_count)
    replace_values['<PRESHARED_KEY>'] = preshared_key
    replace_values['<CS_NAME>'] = cs_name
    replace_values['<WAN1_ISPNAME>'] = wan1_isp_name
    replace_values['<WAN1_CKT_ID>'] = wan1_isp_ckt_id
    replace_values['<WAN1_SUPPORT_NO>'] = wan1_isp_support_num
    replace_values['<WAN1_UP>'] = wan1_isp_up
    replace_values['<WAN1_DOWN>'] = wan1_isp_down
    replace_values['<WAN1_IP>'] = wan1_isp_ip
    replace_values['<WAN1_GW>'] = wan1_isp_gw
    replace_values['<BGP_ASN>'] = bgp_asn
    replace_values['<SITE_ADDRESS>'] = site_mailing_addr
    replace_values['<ATL_TUN1_IP>'] = site_tun1_ip
    replace_values['<ATL_TUN1_PEER>'] = atl_tun1_ip
    replace_values['<DAL_TUN1_IP>'] = site_dal_tun1_ip
    replace_values['<DAL_TUN1_PEER>'] = dal_tun1_ip
    replace_values['<CS_UPLINK_N0>'] = cs_uplink_n0
    replace_values['<CS_UPLINK_N1>'] = cs_uplink_n1
    replace_values['<SITE_SUPERNET>'] = site_supernet[0] + '/' + site_supernet[1]
    replace_values['<SUB>'] = sub
    replace_values['<SITE>'] = site
    replace_values['<TIER>'] = tier
    replace_values['<DEVICE_TYPE>'] = dtype
    replace_values['<DEVICE_SERIAL>'] = dserial
    replace_values['<SWITCH_COUNT>'] = num_switches
    replace_values['<EF01_HOSTNAME>'] = ef_hostname
    replace_values['<CS01_HOSTNAME>'] = cs_name
    replace_values['<TUN1_ST>'] = st_tun



#defaults
    replace_values['<WAN2_ISPNAME>'] = wan2_isp_name
    replace_values['<WAN2_CKT_ID>'] = wan2_isp_ckt_id
    replace_values['<WAN2_SUPPORT_NO>'] = wan2_isp_support_num
    replace_values['<WAN2_UP>'] = wan2_isp_up
    replace_values['<WAN2_DOWN>'] = wan2_isp_down
    replace_values['<WAN2_IP>'] = wan2_isp_ip
    replace_values['<WAN2_GW>'] = wan2_isp_gw
    replace_values['<VC_N0_SN>'] = vc_n0_sn
    replace_values['<VC_N1_SN>'] = vc_n1_sn
    replace_values['<ATL_TUN2_IP>'] = site_atl_tun2_ip
    replace_values['<ATL_TUN2_PEER>'] = atl_tun2_ip
    replace_values['<DAL_TUN2_IP>'] = site_dal_tun2_ip
    replace_values['<DAL_TUN2_PEER>'] = dal_tun2_ip
    replace_values['<TUN2_ST>'] = st_tun

    if tier == 1:

        atl_tun2_ip = input('What is the 2nd ATL Tunnel IP xx.xx.xx.xx : ')
        site_atl_tun2_ip = input('What is the 2nd Site to ATL Tunnel IP xx.xx.xx.xx : ')

        dal_tun2_ip = input('What is the 2nd Dal Tunnel IP xx.xx.xx.xx : ')
        site_dal_tun2_ip = input('what is the 2nd Site to Dal Tunnel IP xx.xx.xx.xx : ')

        wan2_isp_name = input('WAN 2 ISP Name: ')
        wan2_isp_ckt_id = input('WAN 2 ISP CKT ID: ')
        wan2_isp_support_num = input('WAN 2 ISP Support Number: ')
        wan2_isp_up = input('WAN 2 ISP Upload Speed: ')
        wan2_isp_down = input('WAN 2 ISP Download Speed: ')
        wan2_isp_ip = input('WAN 2 ISP IP xx.xx.xx.xx/xx : ' )
        wan2_isp_gw = input('WAN 2 ISP GW xx.xx.xx.xx : ')

        vc_n0_sn = input('VC N0 SN: ')
        vc_n1_sn = input('VC N1 SN: ')
        st_tun2 = input('ST Tunnel #2: ')


        #Add Token Key Pairs to Replace Values
        replace_values['<WAN2_ISPNAME>'] = wan2_isp_name
        replace_values['<WAN2_CKT_ID>'] = wan2_isp_ckt_id
        replace_values['<WAN2_SUPPORT_NO>'] = wan2_isp_support_num
        replace_values['<WAN2_UP>'] = wan2_isp_up
        replace_values['<WAN2_DOWN>'] = wan2_isp_down
        replace_values['<WAN2_IP>'] = wan2_isp_ip
        replace_values['<WAN2_GW>'] = wan2_isp_gw
        replace_values['<<VC_N0_SN>'] = vc_n0_sn
        replace_values['<VC_N1_SN>'] = vc_n1_sn
        replace_values['<ATL_TUN2_IP>'] = site_atl_tun2_ip
        replace_values['<ATL_TUN2_PEER>'] = atl_tun2_ip
        replace_values['<DAL_TUN2_IP>'] = site_dal_tun2_ip
        replace_values['<DAL_TUN2_PEER>'] = dal_tun2_ip
        replace_values['<TUN2_ST>'] = st_tun2

    data = replace_values
# Run SRX Config to Jinja2
    jinsrx(data)
# Run EX Core Config to Jinja2
    hostname = replace_values['<HOSTNAME>']
    core_switch_hostname = hostname[0:len(hostname) - 4] + 'CS01'
    replace_values['<HOSTNAME>'] = core_switch_hostname
    jinex(data)
# Run EX VC Config to Jinja2
    if replace_values['VC'] is not 'None':
        replace_values['ACCESS'] = 'Yes'
        hostname = replace_values['<HOSTNAME>']
        core_switch_hostname = hostname[0:len(hostname) - 4] + 'AS01'
        replace_values['<HOSTNAME>'] = core_switch_hostname
        jinex(data)
# Run Head End Config to Jinja2
    hostname = replace_values['<HOSTNAME>']
    core_switch_hostname = hostname[0:len(hostname) - 4] + 'HE'
    replace_values['<HOSTNAME>'] = core_switch_hostname
    jinhe(data)


#Subnet Breakdown
# Create a Scope for a Tier 1 , Tier 2, Tier 3 Setup
# But prompt user to allow custom sizing if desired.

#Tier 1
#Site Supernet	10.xx.0.0/22	EF Loopback	10.xx.xx(+3).255/32
#VLAN 300	10.xx.0.0/24	CS Loopback	10.xx.xx(+3).254/32
#VLAN 310	10.xx.xx(+1).0/24	AS Loopback	10.xx.xx(+3).253/32  (+256)
#VLAN 320	10.xx.xx(+2).0/24	AS Loopback	10.xx.xx(+3).252/32
#VLAN 330	10.xx.xx(+3).0/25	AS Loopback	10.xx.xx(+3).251/32
#VLAN 311	10.xx.xx(+3).128/26	EF reth0.901 10.xx.xx(+3).192/31
#VLAN 901	10.xx.xx(+3).192/31	CS irb.901	10.xx.xx(+3).193/31
#VLAN 901   10.xx.xx(+1).224/27 Loopbacks .225 - .255

## Tier 2/3
## Site Supernet	10.xx.xx.0/23	EF Loopback	10.xx.xx(+1).255
#VLAN 300	10.xx.xx.0/25	CS Loopback	10.xx.xx(+1).254
#VLAN 310	10.xx.xx.128/25	AS Loopback	10.xx.xx(+1).253
#VLAN 320	10.xx.xx(+1).0/25	AS Loopback	10.xx.xx(+1).254
#VLAN 330	10.xx.xx(+1).128/26	AS Loopback	10.xx.xx(+1).253
#VLAN 311	10.xx.xx(+1).192/27	EF ge-0/0/5.0 10.xx.xx(+1).224/31
#VLAN 901	10.xx.xx(+1).224/31	CS irb.901 10.xx.xx(+1).225/31
#VLAN 901   10.xx.xx(+1).227/27 Loopbacks


# 255.255.255.0
# 11111111.11111111.11111100.000000000
# + 256 = 0

def subnet_cont_tier1(site_supernet):
    CIDR_MASK = 4294966272

    # pprint.pprint(socket.inet_aton(site_supernet[0]))
    site_supernet_num = struct.unpack('>i', socket.inet_aton(site_supernet[0]))[0]

    base_supernet = site_supernet_num & CIDR_MASK
    # print(socket.inet_ntoa(base_supernet))

    vlan300_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 1)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 10)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 254)), 'cidr': '24'}
    vlan310_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 256)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 257)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 263)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 510)), 'cidr': '24'}
    vlan320_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 512)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 513)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 519)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 766)), 'cidr': '24'}
    vlan330_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 768)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 769)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 775)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 894)), 'cidr': '25'}
    vlan311_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 896)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 897)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 900)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 958)), 'cidr': '26'}
    vlan901_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 960)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 960)),
                   'gw': socket.inet_ntoa(struct.pack('>i', base_supernet + 961)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 962)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 1021)),
                   'loopback': socket.inet_ntoa(struct.pack('>i', base_supernet + 555)),
                   'ef_loopback': socket.inet_ntoa(struct.pack('>i', base_supernet + 1023)),
                   'cs_loopback': socket.inet_ntoa(struct.pack('>i', base_supernet + 1022)),
                   'cidr': '26'}

    return {'vlan300': vlan300_net,
            'vlan310': vlan310_net,
            'vlan320': vlan320_net,
            'vlan330': vlan330_net,
            'vlan311': vlan311_net,
            'vlan901': vlan901_net}


def subnet_const_tier3(site_supernet):
    CIDR_MASK = 4294966784

    # pprint.pprint(socket.inet_aton(site_supernet[0]))
    site_supernet_num = struct.unpack('>i', socket.inet_aton(site_supernet[0]))[0]

    base_supernet = site_supernet_num & CIDR_MASK
    # print(socket.inet_ntoa(base_supernet))

    vlan300_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 1)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 10)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 126)), 'cidr': '25'}
    vlan320_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 128)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 129)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 135)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 254)), 'cidr': '25'}
    vlan310_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 256)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 257)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 263)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 382)), 'cidr': '25'}
    vlan311_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 384)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 385)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 386)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 446)), 'cidr': '26'}
    vlan330_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 448)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 449)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 454)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 480)), 'cidr': '27'}
    vlan901_net = {'net': socket.inet_ntoa(struct.pack('>i', base_supernet + 480)),
                   'ip': socket.inet_ntoa(struct.pack('>i', base_supernet + 480)),
                   'gw' : socket.inet_ntoa(struct.pack('>i', base_supernet + 481)),
                   'dhcplow': socket.inet_ntoa(struct.pack('>i', base_supernet + 483)),
                   'dhcphigh': socket.inet_ntoa(struct.pack('>i', base_supernet + 509)),
                   'ef_loopback': socket.inet_ntoa(struct.pack('>i', base_supernet + 511)),
                   'cs_loopback': socket.inet_ntoa(struct.pack('>i', base_supernet + 510)),
                   'cidr': '27'}


    return {'vlan300': vlan300_net,
            'vlan310': vlan310_net,
            'vlan320': vlan320_net,
            'vlan330': vlan330_net,
            'vlan311': vlan311_net,
            'vlan901': vlan901_net}

# ef_Single_Template
# Branch Firewall (300, 320, 340, 1500)
# Slight Variation in different SRX Models
# Only 1 ATL / Dal VPN Tunel
# only 1 connection to switch



#Starting the program


if __name__ == '__main__':
    print('Generating firewall configs for remote sites')
#    getinput()
    print('It all worked, we are done.')
